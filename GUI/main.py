import datetime
from datetime import datetime
from statistics import mean
import time
import sys
from RadialBar import RadialBar
from PyQt5 import QtGui, QtCore, QtGui, QtWidgets, QtQml
from PyQt5.QtCore import pyqtSlot, pyqtSignal, QUrl
from PyQt5.QtQuick   import QQuickView

import paho.mqtt.client as paho
from openpyxl import Workbook, workbook,load_workbook
#broker="123.45.0.10" / test.mosquitto.org
broker="127.0.0.1"
port = 1883

S1 = 0
S2 = 0
S3 = 0
S4 = 0
S5 = 0
S6 = 0

time_now = 0
time_prev = time.time()
time_changesheet = 0
time_prev_changesheet = time.time()
now = time.strftime("%x")
wb1 = Workbook()
ws1 = wb1.active
wb2 = Workbook()
ws2 = wb2.active

def on_message(client, userdata, message):
	msg = str(message.payload.decode("utf-8"))
	t = str(message.topic)

	if(msg[0] == 'c'):
		val =  1
	else:
		val = str(msg)

	if (t == "FlowRate"):
		global S1
		S1 = float(msg)
		
	if (t == "TotalLiter"):
		global S2
		S2 = int(msg)
		
	if (t == "statusesp"):
		global S3
		if (msg == "on"):
			S3 = "ON"
		else:
			S3 = "OFF"

	if (t == "Temp"):
		global S4
		S4 = float(msg)

	if (t == "Humidity"):
		global S5
		S5 = float(msg)

	if (t == "DC Motor Slider CW1"):
		global S6
		S6 = int(msg)

	if (t == "Temp"):
		print(f"Suhu : {S4} --- Humiditas : {S5}")
		print(f"Laju Aliran : {S1} --- Total Liter : {S2}")
		print(f"Nilai PWM DC Motor : {S6}")


class Board(QQuickView):  

	####### NOTES FOR TRANSFER FROM PYTHON TO QML #######
	FlowRate = pyqtSignal(str)
	TotalLiter = pyqtSignal(str)
	statusesp = pyqtSignal(str)
	Temp = pyqtSignal(str)
	Humidity = pyqtSignal(str)
	DCSliderCW1 = pyqtSignal(str)

	
	

	def __init__(self):
		super().__init__()
		self.setSource(QUrl('mainqml.qml'))
		self.rootContext().setContextProperty("Board", self) 
		self.setFlags(QtCore.Qt.FramelessWindowHint | QtCore.Qt.Window | QtCore.Qt.WindowStaysOnTopHint)
		self.setResizeMode(QQuickView.SizeRootObjectToView)		
		self.show()
		vista = self.rootObject()
		self.starttimer()
		self.FlowRate.connect(vista.flow_rate)
		self.TotalLiter.connect(vista.total_liter)
		self.statusesp.connect(vista.esp_status)
		self.Temp.connect(vista.temp_status)
		self.Humidity.connect(vista.hum_status)
		
		
		
    ####### DATA UNTUK DITRANSFER DARI QML KE PYTHON ARDUINO #######

	@pyqtSlot(result=int)
	def get_time(self):
		date_time = datetime.currentDateTime()
		unixTIME = date_time.toSecsSinceEpoch()
		#unixTIMEx = date_time.currentMSecsSinceEpoch()
		return unixTIME

	@pyqtSlot('QString')
	def setquit(self, value):
		global S2
		dataquit = str(value)
		print(dataquit)
		if dataquit == "Quit Now" :
			S2=0
			self.close()

			

	@pyqtSlot('QString') 
	def setmax(self, value):
		datamax = str(value)
		print(datamax)
		if datamax == "Maximize Now" :
			self.showMaximized()
		else : 
			self.show()

	@pyqtSlot('QString')
	def setmini(self, value):
		datamini = str(value)
		print(datamini)
		if datamini == "Minimize Now" :
			self.showMinimized()

	@pyqtSlot('QString')
	def setslidercw1(self,value):
		dataslidercw1 =value
		client.publish("DC Motor Slider CW1", dataslidercw1)
		print (dataslidercw1)

	@pyqtSlot('QString')
	def setsliderccw1(self,value):
		datasliderccw1 =value
		client.publish("DC Motor Slider CCW1", datasliderccw1)
		print (datasliderccw1)

	@pyqtSlot('QString')
	def setslidercw2(self,value):
		dataslidercw2 =value
		client.publish("DC Motor Slider CW2", dataslidercw2)
		print (dataslidercw2)

	@pyqtSlot('QString')
	def setsliderccw2(self,value):
		datasliderccw2 =value
		client.publish("DC Motor Slider CCW2", datasliderccw2)
		print (datasliderccw2)
	

	def starttimer(self):
		self.timer = QtCore.QTimer()
		self.timer.timeout.connect(self.sampling)
		self.timer.start(500)
		
	def sampling(self):
		####### TRANSFER TO QML #######
		global time_prev
		global time_now
		global time_changesheet
		global time_prev_changesheet
		global ws1
		global ws2
		global now
		st1=S1
		st2=S2
		st3=S3
		st4=S4
		st5=S5
		st6=S6

			######## SEND DATA TO TRANSFER FUNCTION ########
		self.FlowRate.emit(str(st1))
		self.TotalLiter.emit(str(st2))
		self.statusesp.emit(str(st3))
		self.Temp.emit(str(st4))
		self.Humidity.emit(str(st5))
		self.DCSliderCW1.emit(str(st6))
		time_now = time.time() - time_prev
		time_changesheet = time.time() - time_prev_changesheet
		ws1['A1'] = "Tanggal"
		ws1['B1'] = "Waktu"
		ws1['C1'] = "Temp(°C)"
		ws1['D1'] = "Hum(RH)"

		ws2['A1'] = "Tanggal"
		ws2['B1'] = "Waktu"
		ws2['C1'] = "Laju Air (L/M)"
		ws2['F1'] = "Total (ML)"
		ws2['F2'] = st2

		if (time_changesheet > 62):
			time_prev_changesheet = time.time()
			ws1 = wb1.create_sheet("TempHump")
			ws2 = wb2.create_sheet("WaterFlow")
			wb1.active = ws1
			wb2.active = ws2

		if (time_now > 1):
			time_prev = time.time()
			clock = datetime.now()
			current_time = clock.strftime("%H:%M:%S")
			ws1 = wb1.active
			ws2 = wb2.active
			ws1.append([now,current_time,st4, st5]) #insert value to excel
			ws2.append([now,current_time,st1])
			#wb1.create_sheet("Hari") #create new sheet
			wb1.save("TempHump.xlsx")
			wb2.save("WaterFlow.xlsx")
		
		print(time_changesheet)

		self.timer.start(50)
		

if __name__ == '__main__':
    
	app = QtWidgets.QApplication(sys.argv)
	
	QtQml.qmlRegisterType(RadialBar, "SDK", 1,0, "RadialBar")

	### MQTT ###
	print("Creating new instance")
	client= paho.Client("GUI")
	client.on_message=on_message

	print("connecting to broker ",broker)
	client.connect(broker,port)#connect
	print(broker," connected")


	client.loop_start()
	print("Subscribing to topic")
	
	client.subscribe("FlowRate")
	client.subscribe("TotalLiter")
	client.subscribe("statusesp")
	client.subscribe("Temp")
	client.subscribe("Humidity")

	

	### Connection ##
	w = Board()
	sys.exit(app.exec_())


	